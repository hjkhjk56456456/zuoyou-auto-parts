#!/usr/bin/env python3
# 合并 498 个 xlsx -> 50 个分片 JSON + 完整 Excel + 索引
import openpyxl, glob, os, json, re

SRC = r'C:\Users\admin\Desktop\济南佐佑_530行拆分_498个文档'
OUT = r'C:\Users\admin\github\zuoyou-auto-parts'
PART_DIR = os.path.join(OUT, 'catalog', 'parts')
DL_DIR = os.path.join(OUT, 'downloads')
os.makedirs(PART_DIR, exist_ok=True)
os.makedirs(DL_DIR, exist_ok=True)

files = sorted([f for f in glob.glob(os.path.join(SRC, '*.xlsx')) if not os.path.basename(f).startswith('~')],
               key=lambda p: int(re.match(r'(\d+)', os.path.basename(p)).group(1)))
print(f'读取 {len(files)} 个文件...')

seen = set()
items = []
for fi, f in enumerate(files, 1):
    wb = openpyxl.load_workbook(f, read_only=True)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        cell = row[0]
        if not cell:
            continue
        s = str(cell).strip()
        if not s:
            continue
        # 解析: 'OEM型号 Zuoyou, 佐佑型号1 Zuoyou, 佐佑型号2 ...'
        parts = [p.strip() for p in s.split('Zuoyou,')]
        oem = parts[0]
        # 清洗每段末尾可能残留的 'Zuoyou'（无逗号结尾的脏数据）
        zy = ' / '.join(p.rstrip('Zuoyou').strip() for p in parts[1:] if p.strip())
        if not oem:
            continue
        key = (oem.lower(), zy.lower())
        if key in seen:
            continue
        seen.add(key)
        items.append({'o': oem, 'z': zy, 'f': fi})
    wb.close()
    if fi % 100 == 0:
        print(f'  已处理 {fi}/{len(files)} 文件, 累计 {len(items)} 条')

# 排序：数字优先，再按字母
def sort_key(it):
    o = it['o']
    return (0, o.zfill(20)) if o.isdigit() else (1, o.lower())
items.sort(key=sort_key)

total = len(items)
print(f'去重后总条数: {total}')

# 分片: 目标 50 片
N_PARTS = 50
per = max(1, total // N_PARTS)
parts_meta = []
for i in range(N_PARTS):
    chunk = items[i*per:(i+1)*per] if i < N_PARTS-1 else items[i*per:]
    if not chunk:
        break
    name = f'part-{i+1:03d}.json'
    with open(os.path.join(PART_DIR, name), 'w', encoding='utf-8') as fh:
        json.dump([{'o': it['o'], 'z': it['z']} for it in chunk], fh, ensure_ascii=False, separators=(',', ':'))
    parts_meta.append(name)
    print(f'  分片 {name}: {len(chunk)} 条')

with open(os.path.join(OUT, 'catalog', 'index.json'), 'w', encoding='utf-8') as fh:
    json.dump({'total': total, 'parts': len(parts_meta), 'files': parts_meta}, fh, ensure_ascii=False)
print(f'索引完成: {len(parts_meta)} 片')

# 完整 Excel 下载版
print('生成完整 Excel...')
wb = openpyxl.Workbook(write_only=True)
ws = wb.create_sheet('产品目录')
ws.append(['序号', 'OEM原厂型号', '佐佑型号', '来源文档编号'])
for i, it in enumerate(items, 1):
    ws.append([i, it['o'], it['z'], it['f']])
xl = os.path.join(DL_DIR, '佐佑产品目录_完整版.xlsx')
wb.save(xl)
print(f'Excel 完成: {os.path.getsize(xl)/1024/1024:.1f} MB -> {xl}')

# 随机样本（首页秒开展示用）
import random
random.seed(42)
sample = random.sample(items, min(200, len(items)))
random.shuffle(sample)
with open(os.path.join(OUT, 'catalog', 'sample.json'), 'w', encoding='utf-8') as fh:
    json.dump([{'o': it['o'], 'z': it['z']} for it in sample], fh, ensure_ascii=False, separators=(',', ':'))
print(f'样本: {len(sample)} 条 -> catalog/sample.json')
print('DONE')
